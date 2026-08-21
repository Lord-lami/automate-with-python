import sys, openpyxl
from openpyxl.styles import Font

# Creates a W x H multiplication table in an Excel file

# Takes the width and height of the table from the command line
width = 0
height = 0
if len(sys.argv) > 1:
    width = int(sys.argv[1])
else:
    print("No width or height given", file=sys.stderr)

if len(sys.argv) > 2:
    height = int(sys.argv[2])
else:
    height = width

# Creates a new Workbook object
wb = openpyxl.Workbook()
sheet = wb.active

# Creates the labels in bold red on row 1 and column A
bold_red = Font(bold=True, color="FFFF0000")
for w in range(1, width+1):
    sheet.cell(1, w+1, w).font = bold_red
for h in range(1, height+1):
    sheet.cell(h+1, 1, h).font = bold_red

# Creates the table in between row 1 and column A
for w in range(1, width+1):
    for h in range(1, height+1):
        sheet.cell(h+1, w+1, h * w)

# Saves the Workbook object to multiplicationTable.xlsx
wb.save('multiplicationTable.xlsx')
